#!/usr/bin/env python3
"""
BILAN DE PUISSANCE prévisionnel (avant-projet) — Polyclinique Nassib.
Modèle transparent à partir de docs/fiches-locaux/rooms-data.json :
  charges par local (éclairage, petite force, équipement biomédical, CVC, informatique)
  → agrégation par tableau divisionnaire de zone (TD) → coefficients ku/ks
  → puissance d'utilisation, répartie Normal / Secouru (groupe) / Ondulé (ASI)
  → dimensionnement TGBT / transformateur / groupe / onduleur.

Toutes les valeurs unitaires sont des HYPOTHÈSES d'avant-projet (feuille "Hypothèses",
éditable) — à valider et recalculer par le bureau d'études électricité.

Sortie : docs/fiches-locaux/bilan-puissance-nassib.xlsx
         docs/BILAN_PUISSANCE.md (synthèse)
"""
import os, json, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = json.load(open(os.path.join(BASE, "docs", "fiches-locaux", "rooms-data.json")))
XLSX = os.path.join(BASE, "docs", "fiches-locaux", "bilan-puissance-nassib.xlsx")
MD = os.path.join(BASE, "docs", "BILAN_PUISSANCE.md")

# ───────── Hypothèses unitaires (éditables) ─────────
LIGHT_WM2 = {  # éclairage W/m²
    "bloc_cesarienne": 24, "bloc_operatoire": 24, "sspi_reveil": 22, "sterilisation": 24,
    "laboratoire": 24, "imagerie": 20, "urgences_box": 18, "urgences_dechocage": 18,
    "salle_travail": 18, "chambre_patient": 18, "hospit_jour": 18, "biberonnerie": 18,
    "consultation": 15, "accueil_admin": 12, "support": 8, "locaux_techniques_vrd": 10,
}
EQUIP_KW = {  # équipement biomédical / force spécifique (kW installé) par local type
    "imagerie": 30, "sterilisation": 45, "laboratoire": 8, "bloc_cesarienne": 8,
    "bloc_operatoire": 6, "sspi_reveil": 4, "urgences_dechocage": 5, "urgences_box": 1.5,
    "chambre_patient": 0.6, "hospit_jour": 2, "biberonnerie": 3, "salle_travail": 2,
    "consultation": 1.0, "accueil_admin": 0.5, "support": 0.2, "locaux_techniques_vrd": 0,
}
KW_PER_PC = 0.10      # petite force : kW installé foisonné par prise 16A
KW_PER_RJ = 0.10      # informatique (ondulé) : kW par RJ45
EER = 3.0             # rendement froid : kW élec = kW froid / EER
COSPHI = 0.90
KU = {"light": 1.0, "prises": 0.25, "equip": 0.6, "cvc": 0.85, "info": 0.6, "tech": 0.75}
KS_TD = 0.85          # simultanéité au niveau d'un TD
KS_GLOBAL = 0.90      # simultanéité au TGBT
EXTENSION = 0.20      # réserve d'extension (20 %)
MARGE_GROUPE = 1.25
MARGE_ASI = 1.30

# Charges spécifiques du local technique extérieur (TEC-01), kW installé
TEC_LOADS = [  # (désignation, kW, categorie)
    ("Production O₂ (PSA/concentrateur)", 15, "tech"),
    ("Centrale air médical (compresseurs)", 15, "tech"),
    ("Centrale de vide médical (pompes)", 10, "tech"),
    ("Ascenseur", 12, "tech"),
    ("Centrale froid / CTA communes", 30, "cvc"),
    ("Services généraux · ECS · VRD", 25, "general"),
]

# ───────── Mapping local → tableau de zone (TD) ─────────
def td_of(r):
    c = r["code"]
    if c == "TEC-01":
        return "TEC-01 — Local technique"
    if c.startswith("MAT"):
        return "R1 · TD-HOSPI MATERNITÉ"
    if c.startswith("HOS") or c == "HDJ-01":
        return "R1 · TD-HOSPI MÉDECINE"
    if c == "BIB-01" or c == "INF-R1":
        return "R1 · TD-NÉONAT/INFIRMERIE"
    if c.startswith("BUR-R1"):
        return "R1 · TD-BUREAUX"
    if c in ("ACC-R1", "ATT-R1", "ADM-R1"):
        return "R1 · TD-ADMIN"
    if c in ("BLC-01", "SAS-BLC", "REV-01", "STE-01"):
        return "RDC · TD-BLOC/SSPI/STÉRIL"
    if c in ("IMG-01", "SAS-IMG", "LAB-01", "MLAB-01"):
        return "RDC · TD-IMAGERIE/LABO"
    if c in ("PHA-01", "MAG-01", "VES-H", "VES-F"):
        return "RDC · TD-PHARMACIE/SUPPORT"
    if c in ("ACC-01", "ADM-01", "CAI-01", "ATT-01"):
        return "RDC · TD-ACCUEIL/ADMIN"
    if c in ("BCS-01", "BCS-02", "BCS-03", "BCS-04", "DEN-01", "BGYN-01", "BGYN-02"):
        return "RDC · TD-CONSULT/GYN"
    if r["zone"] == "gyneco_obstetrique":   # pré-travail, travail, infirmerie/bureaux mat.
        return "RDC · TD-MATERNITÉ (plateau)"
    if r["zone"] == "urgences":
        return "RDC · TD-URGENCES (IT médical)"
    return "RDC · TD-DIVERS"

# Répartition (normal, secouru, ondulé) par type de charge et criticité
def split(cat, crit):
    if cat == "info":
        return (0, 0, 1.0)
    if cat == "tech":
        return (0, 1.0, 0)
    if cat == "general":
        return (1.0, 0, 0)
    if cat == "light":
        return (0, 1.0, 0) if crit in ("C", "E") else (0.6, 0.4, 0)
    if cat == "prises":
        return {"C": (0, 1.0, 0), "E": (0.5, 0.5, 0)}.get(crit, (1.0, 0, 0))
    if cat == "equip":
        return (0, 0.7, 0.3) if crit in ("C", "E") else (0.3, 0.7, 0)
    if cat == "cvc":
        return {"C": (0, 1.0, 0), "E": (0.5, 0.5, 0)}.get(crit, (1.0, 0, 0))
    return (1.0, 0, 0)

# ───────── Calcul par local ─────────
def room_loads(r):
    t = r["template"]
    pi = {
        "light": r["area"] * LIGHT_WM2.get(t, 12) / 1000.0,
        "prises": r["pc16"] * KW_PER_PC,
        "equip": EQUIP_KW.get(t, 0.5),
        "cvc": (r["coolKw"] or 0) / EER,
        "info": r["rj45"] * KW_PER_RJ,
    }
    if r["code"] == "PHA-01":
        pi["equip"] += 2.5          # enceintes réfrigérées
    if r["code"] in ("BGYN-01", "BGYN-02"):
        pi["equip"] += 0.5          # échographe
    return pi

# Agrégation par TD
TD = {}
for r in DATA:
    if r["code"] == "TEC-01":
        continue
    name = td_of(r)
    td = TD.setdefault(name, {"pi": {}, "pu": {}, "nso": [0, 0, 0], "crit": r["crit"]})
    for cat, p in room_loads(r).items():
        pu = p * KU[cat]
        td["pi"][cat] = td["pi"].get(cat, 0) + p
        td["pu"][cat] = td["pu"].get(cat, 0) + pu
        n, s, o = split(cat, r["crit"])
        td["nso"][0] += pu * n; td["nso"][1] += pu * s; td["nso"][2] += pu * o

# TEC-01
tec = {"pi": {}, "pu": {}, "nso": [0, 0, 0], "crit": "C"}
for label, kw, cat in TEC_LOADS:
    pu = kw * KU.get(cat, 0.75)
    tec["pi"][cat] = tec["pi"].get(cat, 0) + kw
    tec["pu"][cat] = tec["pu"].get(cat, 0) + pu
    n, s, o = split(cat, "C")
    tec["nso"][0] += pu * n; tec["nso"][1] += pu * s; tec["nso"][2] += pu * o
TD["TEC-01 — Local technique"] = tec

# Application ks_TD
rows = []
for name, td in TD.items():
    pi_tot = sum(td["pi"].values())
    pu_tot = sum(td["pu"].values()) * KS_TD
    n, s, o = [v * KS_TD for v in td["nso"]]
    rows.append({"td": name, "pi": pi_tot, "pu": pu_tot, "n": n, "s": s, "o": o, "crit": td["crit"],
                 "cat": {k: td["pi"].get(k, 0) for k in ("light", "prises", "equip", "cvc", "info", "tech", "general")}})

rows.sort(key=lambda x: (0 if x["td"].startswith("RDC") else 1 if x["td"].startswith("R1") else 2, x["td"]))

# Totaux globaux
Pi = sum(r["pi"] for r in rows)
Pu = sum(r["pu"] for r in rows) * KS_GLOBAL
PN = sum(r["n"] for r in rows) * KS_GLOBAL
PS = sum(r["s"] for r in rows) * KS_GLOBAL
PO = sum(r["o"] for r in rows) * KS_GLOBAL
S_kva = Pu / COSPHI
S_ext = S_kva * (1 + EXTENSION)
# dimensionnement
def std_transfo(s):
    for v in (160, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000):
        if v >= s:
            return v
    return 2500
transfo = std_transfo(S_ext)
groupe = PS * MARGE_GROUPE / COSPHI
def std_groupe(s):
    for v in (100, 150, 200, 250, 300, 400, 500, 630, 800):
        if v >= s:
            return v
    return 1000
groupe_std = std_groupe(groupe)
asi = PO * MARGE_ASI / COSPHI

# ───────── Excel ─────────
wb = Workbook()
thin = Side(style="thin", color="BFC7CE")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
C = Alignment("center", "center", wrap_text=True)
L = Alignment("left", "center", wrap_text=True)
def hdr(cell, txt, color="1F4E79"):
    cell.value = txt; cell.font = Font(bold=True, color="FFFFFF", size=9)
    cell.fill = PatternFill("solid", fgColor=color); cell.alignment = C; cell.border = bd

ws = wb.active; ws.title = "Bilan par TD"
ws.merge_cells("A1:L1")
t = ws["A1"]; t.value = "BILAN DE PUISSANCE PRÉVISIONNEL — Polyclinique Nassib (FIOG) · avant-projet — à valider par le BE électricité"
t.font = Font(bold=True, color="FFFFFF", size=12); t.fill = PatternFill("solid", fgColor="1F4E79"); t.alignment = L
ws.row_dimensions[1].height = 22
heads = ["Tableau de zone (TD)", "Crit.", "Pi éclairage", "Pi prises", "Pi équip.", "Pi CVC",
         "Pi info", "Pi techn.", "Pi total kW", "Pu kW", "dont Secouru", "dont Ondulé"]
for i, h in enumerate(heads, 1):
    hdr(ws.cell(2, i), h)
ws.row_dimensions[2].height = 32
ri = 3
CF = {"C": "F4CCCC", "E": "FCE5CD", "M": "FFF2CC", "F": "D9EAD3"}
for r in rows:
    cat = r["cat"]
    vals = [r["td"], r["crit"],
            round(cat["light"], 1), round(cat["prises"], 1), round(cat["equip"], 1),
            round(cat["cvc"], 1), round(cat["info"], 1), round(cat["tech"] + cat["general"], 1),
            round(r["pi"], 1), round(r["pu"], 1), round(r["s"], 1), round(r["o"], 1)]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(ri, ci, v); c.border = bd
        c.alignment = L if ci == 1 else C; c.font = Font(size=9, bold=(ci == 1))
        if ci == 2:
            c.fill = PatternFill("solid", fgColor=CF.get(v, "FFFFFF"))
    ri += 1
# ligne total
tot = ["TOTAL (× ks global %.2f)" % KS_GLOBAL, "",
       round(sum(x["cat"]["light"] for x in rows), 1), round(sum(x["cat"]["prises"] for x in rows), 1),
       round(sum(x["cat"]["equip"] for x in rows), 1), round(sum(x["cat"]["cvc"] for x in rows), 1),
       round(sum(x["cat"]["info"] for x in rows), 1), round(sum(x["cat"]["tech"] + x["cat"]["general"] for x in rows), 1),
       round(Pi, 1), round(Pu, 1), round(PS, 1), round(PO, 1)]
for ci, v in enumerate(tot, 1):
    c = ws.cell(ri, ci, v); c.border = bd; c.alignment = L if ci == 1 else C
    c.font = Font(bold=True, size=9, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F4E79")
widths = [34, 7, 11, 10, 10, 9, 9, 10, 11, 10, 11, 11]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:L{ri}"

# Feuille Synthèse
sy = wb.create_sheet("Synthèse & dimensionnement")
sy.column_dimensions["A"].width = 46; sy.column_dimensions["B"].width = 18; sy.column_dimensions["C"].width = 40
def line(r, a, b, c="", bold=False, fill=None):
    sy.cell(r, 1, a).font = Font(bold=bold, size=10)
    cb = sy.cell(r, 2, b); cb.font = Font(bold=True, size=11); cb.alignment = C
    sy.cell(r, 3, c).font = Font(size=9, italic=True)
    if fill:
        for col in (1, 2, 3):
            sy.cell(r, col).fill = PatternFill("solid", fgColor=fill)
sy.merge_cells("A1:C1"); sy["A1"].value = "SYNTHÈSE & DIMENSIONNEMENT"
sy["A1"].font = Font(bold=True, color="FFFFFF", size=12); sy["A1"].fill = PatternFill("solid", fgColor="1F4E79")
rr = 3
line(rr, "Puissance installée totale (Pi)", f"{Pi:.0f} kW", "Somme des charges installées"); rr += 1
line(rr, "Puissance d'utilisation (Pu)", f"{Pu:.0f} kW", f"après ku par charge, ks TD={KS_TD}, ks global={KS_GLOBAL}", bold=True, fill="E2EFDA"); rr += 1
line(rr, "  · dont réseau Normal", f"{PN:.0f} kW", ""); rr += 1
line(rr, "  · dont réseau Secouru (groupe)", f"{PS:.0f} kW", "continuité — actes à risque vital", fill="FCE5CD"); rr += 1
line(rr, "  · dont réseau Ondulé (ASI)", f"{PO:.0f} kW", "informatique, monitorage, têtes de bloc", fill="DDEBF7"); rr += 1
rr += 1
line(rr, f"Puissance apparente (cos φ={COSPHI})", f"{S_kva:.0f} kVA", "S = Pu / cos φ"); rr += 1
line(rr, f"+ réserve d'extension {int(EXTENSION*100)} %", f"{S_ext:.0f} kVA", ""); rr += 1
line(rr, "TRANSFORMATEUR / TGBT — calibre retenu", f"{transfo} kVA", "valeur normalisée ≥ S+extension", bold=True, fill="FFF2CC"); rr += 1
rr += 1
line(rr, "GROUPE ÉLECTROGÈNE", f"≈ {groupe:.0f} kVA", f"Pu secouru × {MARGE_GROUPE} / cos φ"); rr += 1
line(rr, "  · calibre normalisé conseillé", f"{groupe_std} kVA", "à confirmer (démarrages moteurs, ascenseur)", bold=True, fill="FCE5CD"); rr += 1
line(rr, "ONDULEUR CENTRAL (ASI)", f"≈ {asi:.0f} kVA", f"Pu ondulé × {MARGE_ASI} / cos φ — autonomie 10–30 min", bold=True, fill="DDEBF7"); rr += 1
rr += 2
sy.cell(rr, 1, "⚠ Bilan prévisionnel d'avant-projet. Hypothèses unitaires en feuille « Hypothèses ».").font = Font(italic=True, color="C00000", size=9); rr += 1
sy.cell(rr, 1, "À recalculer par le BE (note de calcul) avec la liste d'équipements et les puissances réelles.").font = Font(italic=True, color="808080", size=9)

# Feuille Hypothèses
hp = wb.create_sheet("Hypothèses")
hp.column_dimensions["A"].width = 40; hp.column_dimensions["B"].width = 16; hp.column_dimensions["C"].width = 30
hp.cell(1, 1, "HYPOTHÈSES UNITAIRES (éditables)").font = Font(bold=True, size=12)
hyp = [("Éclairage (W/m²)", "voir détail", "par type de local"),
       ("Petite force — kW/prise 16A", KW_PER_PC, "foisonné"),
       ("Informatique — kW/RJ45", KW_PER_RJ, "ondulé"),
       ("Rendement froid EER", EER, "kW élec = kW froid / EER"),
       ("cos φ", COSPHI, ""),
       ("ku éclairage / prises / équip / CVC / info / techn.", f"{KU['light']}/{KU['prises']}/{KU['equip']}/{KU['cvc']}/{KU['info']}/{KU['tech']}", "coefficients d'utilisation"),
       ("ks au niveau TD", KS_TD, "simultanéité"),
       ("ks global TGBT", KS_GLOBAL, "simultanéité"),
       ("Réserve d'extension", f"{int(EXTENSION*100)} %", ""),
       ("Marge groupe / ASI", f"{MARGE_GROUPE} / {MARGE_ASI}", "")]
for i, (a, b, c) in enumerate(hyp, start=3):
    hp.cell(i, 1, a).font = Font(size=10, bold=True); hp.cell(i, 2, b).alignment = C; hp.cell(i, 3, c).font = Font(size=9, italic=True)
r0 = 3 + len(hyp) + 1
hp.cell(r0, 1, "Équipement biomédical / force spécifique (kW installé) :").font = Font(bold=True)
for i, (k, v) in enumerate(sorted(EQUIP_KW.items()), start=r0 + 1):
    hp.cell(i, 1, k).font = Font(size=9); hp.cell(i, 2, v).alignment = C
r1 = r0 + 1 + len(EQUIP_KW) + 1
hp.cell(r1, 1, "Charges local technique TEC-01 (kW) :").font = Font(bold=True)
for i, (lab, kw, cat) in enumerate(TEC_LOADS, start=r1 + 1):
    hp.cell(i, 1, lab).font = Font(size=9); hp.cell(i, 2, kw).alignment = C; hp.cell(i, 3, cat)

wb.save(XLSX)

# ───────── Markdown synthèse ─────────
md = f"""# Bilan de puissance prévisionnel — Polyclinique Nassib (FIOG)

> **Avant-projet.** Bilan estimatif établi à partir des charges par local (76 locaux) et
> d'hypothèses unitaires (voir feuille « Hypothèses » du classeur). **À recalculer par le
> bureau d'études électricité** (note de calcul) avec la liste d'équipements définitive.
> Départ : TGBT au local technique extérieur arrière ; colonne montante à l'escalier central.

## Synthèse

| Grandeur | Valeur | Commentaire |
|---|---|---|
| Puissance installée (Pi) | **{Pi:.0f} kW** | somme des charges installées |
| Puissance d'utilisation (Pu) | **{Pu:.0f} kW** | après ku/ks (ks TD={KS_TD}, ks global={KS_GLOBAL}) |
| — dont Normal | {PN:.0f} kW | réseau normal |
| — dont Secouru (groupe) | **{PS:.0f} kW** | continuité — actes à risque vital |
| — dont Ondulé (ASI) | **{PO:.0f} kW** | informatique, monitorage, têtes de bloc |
| Puissance apparente | {S_kva:.0f} kVA | S = Pu / cos φ ({COSPHI}) |
| + réserve extension {int(EXTENSION*100)} % | {S_ext:.0f} kVA | |
| **Transformateur / TGBT** | **{transfo} kVA** | calibre normalisé ≥ S+extension |
| **Groupe électrogène** | **{groupe_std} kVA** | (≈ {groupe:.0f} kVA calculé) Pu secouru ×{MARGE_GROUPE} |
| **Onduleur central (ASI)** | **≈ {asi:.0f} kVA** | Pu ondulé ×{MARGE_ASI}, autonomie 10–30 min |

## Répartition par tableau de zone (TD)

| TD | Crit | Pi kW | Pu kW | Secouru kW | Ondulé kW |
|---|---|---|---|---|---|
"""
for r in rows:
    md += f"| {r['td']} | {r['crit']} | {r['pi']:.1f} | {r['pu']:.1f} | {r['s']:.1f} | {r['o']:.1f} |\n"
md += f"| **TOTAL** (ks global {KS_GLOBAL}) | | **{Pi:.0f}** | **{Pu:.0f}** | **{PS:.0f}** | **{PO:.0f}** |\n"
md += """
## Méthode
1. **Charges par local** : éclairage (W/m²), petite force (kW/prise), équipement biomédical
   (kW/type), CVC (kW froid / EER), informatique (kW/RJ45).
2. **ku** (utilisation) par type de charge, puis somme par TD × **ks TD**, puis × **ks global** au TGBT.
3. **Répartition Normal / Secouru / Ondulé** selon le type de charge et la criticité du local
   (locaux critiques et vitaux → secouru ; monitorage/têtes de bloc/informatique → ondulé).
4. **Dimensionnement** : transformateur ≥ S+extension ; groupe = Pu secouru ×1,25 ; ASI = Pu ondulé ×1,30.

## Réserves
- Les puissances d'équipements lourds (RX, autoclave, centrales fluides) sont des **ordres de
  grandeur** — à remplacer par les fiches techniques réelles (R-14, R-08).
- Démarrages moteurs (ascenseur, compresseurs, groupes froid) : vérifier le **régime transitoire**
  du groupe (à-coups de démarrage).
- cos φ et compensation d'énergie réactive à étudier (batterie de condensateurs).
- Sélectivité, chutes de tension et bilan par départ : note de calcul BE.
"""
open(MD, "w").write(md)
print(f"XLSX : {XLSX}")
print(f"MD   : {MD}")
print(f"Pi={Pi:.0f} kW · Pu={Pu:.0f} kW · S={S_kva:.0f} kVA · Transfo={transfo} kVA · Groupe~{groupe_std} kVA · ASI~{asi:.0f} kVA")
print(f"Normal={PN:.0f} · Secouru={PS:.0f} · Ondulé={PO:.0f} kW")
