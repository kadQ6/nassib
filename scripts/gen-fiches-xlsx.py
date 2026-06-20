#!/usr/bin/env python3
"""
Génère le classeur Excel ÉDITABLE des fiches locaux (room-by-room).
Source : docs/fiches-locaux/rooms-data.json (gen-fiches-locaux.mjs).
Sortie : docs/fiches-locaux/fiches-locaux-nassib.xlsx

Feuilles :
  - "Synthèse"     : 1 ligne par local, colonnes groupées (identité, finitions,
                     portes, CVC, CFO, CFA, fluides, plomberie). Filtres + figés.
  - "Légende"      : codes criticité / templates / notes.
"""
import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = json.load(open(os.path.join(BASE, "docs", "fiches-locaux", "rooms-data.json")))
OUT = os.path.join(BASE, "docs", "fiches-locaux", "fiches-locaux-nassib.xlsx")

CRITFILL = {"C": "F4CCCC", "E": "FCE5CD", "M": "FFF2CC", "F": "D9EAD3"}
GRP = {"id": "1F4E79", "fin": "548235", "porte": "7F6000", "cvc": "0F7B6C",
       "cfo": "203864", "cfa": "1155CC", "gaz": "00785A", "plomb": "8A5A00"}

# (clé json, en-tête, groupe, largeur)
COLS = [
    ("code", "Code", "id", 11), ("name", "Intitulé du local", "id", 26),
    ("level", "Niveau", "id", 8), ("sheet", "Plan", "id", 8),
    ("dept", "Service", "id", 16), ("crit", "Criticité", "id", 9),
    ("area", "Surface m²", "id", 10), ("ceilingH", "H. s/FP m", "id", 9), ("volume", "Volume m³", "id", 10),
    ("regime", "Régime électrique", "id", 30),
    # finitions
    ("floor", "Sol", "fin", 32), ("walls", "Murs", "fin", 34), ("ceiling", "Plafond", "fin", 30),
    ("skirt", "Plinthe", "fin", 16), ("upec", "UPEC", "fin", 11), ("chargeSol", "Charge sol daN/m²", "fin", 13),
    # portes
    ("doorW", "Porte larg. mm", "porte", 11), ("doorH", "Porte haut. mm", "porte", 11),
    ("doorColor", "Vantail", "porte", 18), ("doorFrame", "Huisserie", "porte", 16),
    ("doorFire", "Coupe-feu", "porte", 10), ("accessCtrl", "Accès", "porte", 10),
    # CVC
    ("ach", "Renouv. vol/h", "cvc", 11), ("tempC", "T° été °C", "cvc", 9),
    ("humid", "HR %", "cvc", 7), ("surpression", "Surpression Pa", "cvc", 11),
    ("filtration", "Filtration", "cvc", 13), ("coolKw", "Froid kW", "cvc", 9), ("ctaZone", "Zone CTA", "cvc", 16),
    # CFO
    ("pc16", "PC 16A norm.", "cfo", 11), ("ondule", "PC ondulée/sec.", "cfo", 13),
    ("pc20", "PC 20A", "cfo", 8), ("ded", "Alim. dédiée", "cfo", 11),
    ("light", "Éclairage pts", "cfo", 11), ("baes", "BAES", "cfo", 7),
    ("earth", "Terre", "cfo", 7), ("ip", "IP", "cfo", 7),
    # CFA
    ("rj45", "RJ45", "cfa", 7), ("nurse", "Appel malade", "cfa", 11), ("wifi", "Wi-Fi", "cfa", 7),
    ("intercom", "Interphone", "cfa", 10), ("cctv", "Vidéo", "cfa", 7),
    ("access", "Contrôle accès", "cfa", 12), ("tv", "TV", "cfa", 6),
    # fluides
    ("o2", "O₂", "gaz", 6), ("air", "Air méd.", "gaz", 8), ("vide", "Vide", "gaz", 7),
    ("n2oNote", "N₂O", "gaz", 18), ("agssNote", "AGSS", "gaz", 18),
    # plomberie
    ("ef", "EF", "plomb", 6), ("ec", "ECS", "plomb", 7), ("eauTraitee", "Eau traitée", "plomb", 11),
    ("siphons", "Siphons sol", "plomb", 10), ("eu", "EU", "plomb", 6), ("ev", "EV", "plomb", 6),
]

wb = Workbook()
ws = wb.active
ws.title = "Synthèse"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Ligne 1 : titre projet
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
c = ws.cell(1, 1, "POLYCLINIQUE NASSIB — FIOG · Fiches locaux (room-by-room) · indice 180526 — document de conception, valeurs à valider par les BE")
c.font = Font(bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F4E79")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 22

# Ligne 2 : bandeau groupe
GRPLABEL = {"id": "IDENTITÉ", "fin": "FINITIONS", "porte": "PORTES", "cvc": "CVC",
            "cfo": "ÉLECTRICITÉ CFO", "cfa": "COURANTS FAIBLES CFA", "gaz": "FLUIDES MÉDICAUX", "plomb": "PLOMBERIE"}
col = 1
seen = []
order = []
for key, _, grp, _ in COLS:
    order.append(grp)
i = 0
while i < len(COLS):
    grp = COLS[i][2]
    j = i
    while j < len(COLS) and COLS[j][2] == grp:
        j += 1
    ws.merge_cells(start_row=2, start_column=i + 1, end_row=2, end_column=j)
    cc = ws.cell(2, i + 1, GRPLABEL[grp])
    cc.font = Font(bold=True, size=10, color="FFFFFF")
    cc.fill = PatternFill("solid", fgColor=GRP[grp])
    cc.alignment = center
    i = j
ws.row_dimensions[2].height = 18

# Ligne 3 : en-têtes colonnes
for idx, (key, label, grp, width) in enumerate(COLS, start=1):
    cell = ws.cell(3, idx, label)
    cell.font = Font(bold=True, size=9, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=GRP[grp])
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[get_column_letter(idx)].width = width
ws.row_dimensions[3].height = 40

# Données
for ri, r in enumerate(sorted(DATA, key=lambda x: (x["level"], x["code"])), start=4):
    for ci, (key, label, grp, width) in enumerate(COLS, start=1):
        val = r.get(key, "")
        cell = ws.cell(ri, ci, val)
        cell.border = border
        cell.alignment = left if width >= 16 else center
        cell.font = Font(size=9)
        if key == "code":
            cell.font = Font(size=9, bold=True)
        if key == "crit":
            cell.fill = PatternFill("solid", fgColor=CRITFILL.get(val, "FFFFFF"))
            cell.font = Font(size=9, bold=True)
    # surligner toute la ligne légèrement selon criticité (colonne criticité déjà colorée)

ws.freeze_panes = "C4"   # fige identité + en-têtes
ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS))}{3 + len(DATA)}"

# ── Feuille Légende ──
lg = wb.create_sheet("Légende")
lg.column_dimensions["A"].width = 18
lg.column_dimensions["B"].width = 70
rows = [
    ("LÉGENDE — Fiches locaux Nassib", ""),
    ("", ""),
    ("Criticité", "C = Critique (IT médical + secouru + ondulé) · E = Élevé · M = Moyen · F = Faible"),
    ("PC ondulée/sec.", "Prises ondulées (ASI) ou secourues (groupe) — bandeaux/postes critiques"),
    ("Appel malade", "Appel infirmière intégré au bandeau de lit (BTDL) le cas échéant"),
    ("Fluides", "O₂ / Air médical 3,5 bar / Vide. N₂O & AGSS au bloc : à confirmer (réserve R-07)"),
    ("EF / ECS", "Eau froide / Eau chaude sanitaire · EU = eaux usées · EV = eaux vannes"),
    ("UPEC", "Classement d'usage des sols (Usure/Poinçonnement/Eau/Chimie)"),
    ("Source données", "Visible plan A-01/A-02 + fiches FIOG ; valeurs CFO/CFA/CVC déduites des"),
    ("", "templates techniques K'BIO — À DIMENSIONNER ET VALIDER PAR LES BUREAUX D'ÉTUDES."),
    ("Maternité", "Aucune chambre maternité au RDC : les 14 chambres sont au R+1 (plan A-02)."),
    ("Avertissement", "Document de conception — ne vaut pas plan d'exécution ni note de calcul."),
]
for i, (a, b) in enumerate(rows, start=1):
    ca = lg.cell(i, 1, a); cb = lg.cell(i, 2, b)
    if i == 1:
        ca.font = Font(bold=True, size=13)
    else:
        ca.font = Font(bold=True, size=10); cb.font = Font(size=10)
        cb.alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT)
print("XLSX écrit :", OUT, "·", len(DATA), "locaux ·", len(COLS), "colonnes")
